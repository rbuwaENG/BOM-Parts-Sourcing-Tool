from __future__ import annotations

import io
from typing import Optional

import pandas as pd


def _coerce_price(value: Optional[str]) -> float:
    if not value:
        return 0.0
    s = str(value)
    for tok in ["Rs", "$", "USD", "LKR", ","]:
        s = s.replace(tok, "")
    try:
        return float(s.strip())
    except Exception:
        return 0.0


def build_budget_excel(bom_df: pd.DataFrame, results_df: pd.DataFrame) -> bytes:
    df = results_df.copy()
    bom_view = bom_df.copy()
    bom_view["_join_key"] = bom_view["Part_Name"].fillna("")
    df["_join_key"] = df["BOM Part Name"].fillna("")
    merged = pd.merge(df, bom_view[["_join_key", "Quantity"]], on="_join_key", how="left")
    merged["Quantity"] = merged["Quantity"].fillna(0).astype(int)

    merged["Unit Price"] = merged["Price"].apply(_coerce_price)
    merged["Total Price"] = merged["Unit Price"] * merged["Quantity"]

    # Reorder columns
    cols = [
        "BOM Part Name",
        "Found Part Name",
        "Supplier",
        "Quantity",
        "Price",
        "Unit Price",
        "Total Price",
        "Purchase Link",
        "Datasheet Link",
        "Image",
    ]
    for c in cols:
        if c not in merged.columns:
            merged[c] = None
    budget_df = merged[cols]

    # Write to Excel with a summary total row
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        budget_df.to_excel(writer, sheet_name="Budget", index=False)
        ws = writer.sheets["Budget"]
        # Add header with version
        ws.insert_rows(1)
        ws["A1"] = "BOM Budget (v1.0)"
        # Bold header row (now row 2)
        from openpyxl.styles import Font, Alignment
        header_font = Font(bold=True)
        for cell in ws[2]:
            cell.font = header_font
        # Autosize columns (basic)
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value)) if cell.value is not None else 0)
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max(12, max_len + 2), 60)
        # Append overall total
        total_row_idx = ws.max_row + 2
        ws.cell(row=total_row_idx, column=1, value="Overall Total Cost")
        # Find Total Price column index
        total_col_idx = None
        for idx, cell in enumerate(ws[2], start=1):
            if str(cell.value).strip().lower() == "total price":
                total_col_idx = idx
                break
        if total_col_idx:
            start = 3  # data starts at row 3 due to inserted title row
            end = ws.max_row
            ws.cell(row=total_row_idx, column=total_col_idx, value=f"=SUM({ws.cell(row=start, column=total_col_idx).coordinate}:{ws.cell(row=end, column=total_col_idx).coordinate})")
            ws.cell(row=total_row_idx, column=total_col_idx).alignment = Alignment(horizontal="right")
    buf.seek(0)
    return buf.read()